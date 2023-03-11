from openpyxl import load_workbook

from elasticsearch import Elasticsearch
from elasticsearch import helpers
from decouple import config
import sys
import json
def read_cell(sheet, cell):
    val = sheet[cell].value
    if val is not None:
        return str(val).strip()
    else:
        return ""


def read_excel(path):
    wb = load_workbook(filename=path)
    ws = wb["Sheet1"]
    es = Elasticsearch([config("ELASTICSEARCH_URL")])
    actions = []

    for r in range(2, ws.max_row):
        row = {}
        if ws["A{}".format(r)].value is not None:
            row["pid"] = r-1
            scientific, author = split_name(read_cell(ws, "A{}".format(r)))
            row["scientificName"] = scientific
            row["author"] = author
            row["familyName"] = read_cell(ws, "B{}".format(r))
            row["localName"] = read_cell(ws, "C{}".format(r))
            row["utilizedPart"] = read_cell(ws, "D{}".format(r))
            row["ailment"] = read_cell(ws, "F{}".format(r))
            row["activeCompound"] = read_cell(ws, "G{}".format(r))
            row["pmid"] = read_cell(ws, "H{}".format(r))
            row["pmAcList"] = json.loads(read_cell(ws, "I{}".format(r)))
            actions.append({
                "_index":"mpdb",
                "_id":r-1,
                "_source":row
            })

    actions = actions[:1]
    helpers.bulk(es, actions=actions)

def split_name(name):
    splitted = name.split(" ")
    scientific = " ".join(splitted[0:2])
    author = " ".join(splitted[2:])

    return scientific, author
if __name__ == "__main__":
    excelPath = sys.argv[1]
    read_excel(excelPath)



